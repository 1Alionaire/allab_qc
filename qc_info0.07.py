import threading
from pathlib import Path
from tkinter import Tk, Button, Label, Text, filedialog, messagebox, END, DISABLED, NORMAL
from tkinter.ttk import Progressbar
from openpyxl import Workbook, load_workbook
import pandas as pd
from datetime import date
import json


class QCProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("QC Processor")
        self.root.geometry("600x400")
        self.root.resizable(False, False)
        
        self.folder_path = None
        
        # Кнопка выбора папки
        self.btn_select = Button(
            root, 
            text="Choose Folder", 
            command=self.select_folder,
            width=20,
            height=2
        )
        self.btn_select.pack(pady=10)
        
        # Метка с выбранной папкой
        self.lbl_folder = Label(root, text="Папка не выбрана", wraplength=550)
        self.lbl_folder.pack(pady=5)
        
        # Кнопка запуска
        self.btn_run = Button(
            root, 
            text="Запустить", 
            command=self.run_processing,
            width=20,
            height=2,
            state=DISABLED
        )
        self.btn_run.pack(pady=10)
        
        # Прогресс-бар
        self.progress = Progressbar(root, length=550, mode='determinate')
        self.progress.pack(pady=10)
        
        # Текущий статус
        self.lbl_status = Label(root, text="", font=("Arial", 10))
        self.lbl_status.pack(pady=5)
        
        # Лог
        self.log = Text(root, height=10, width=70, state=DISABLED)
        self.log.pack(pady=10)
    
    def select_folder(self):
        folder = filedialog.askdirectory(title="Выберите папку с Excel-файлами")
        if folder:
            self.folder_path = folder
            self.lbl_folder.config(text=f"Папка: {folder}")
            self.btn_run.config(state=NORMAL)
    
    def log_message(self, message):
        """Добавляет сообщение в лог"""
        self.log.config(state=NORMAL)
        self.log.insert(END, message + "\n")
        self.log.see(END)
        self.log.config(state=DISABLED)
    
    def update_status(self, message):
        """Обновляет статус"""
        self.lbl_status.config(text=message)
        self.root.update()
    
    def run_processing(self):
        """Запускает обработку в отдельном потоке"""
        self.btn_run.config(state=DISABLED)
        self.btn_select.config(state=DISABLED)
        self.progress['value'] = 0
        
        # Очищаем лог
        self.log.config(state=NORMAL)
        self.log.delete(1.0, END)
        self.log.config(state=DISABLED)
        
        # Запускаем в отдельном потоке, чтобы GUI не зависал
        thread = threading.Thread(target=self.process_files)
        thread.start()
    
    def process_files(self):
        """Основная логика обработки файлов"""
        results = []
        counter = 0
        
        # Ищем все Excel-файлы
        self.update_status("Поиск Excel-файлов...")
        all_files = []
        for ext in ("*.xlsx", "*.xlsm", "*.xls"):
            all_files.extend(Path(self.folder_path).rglob(ext))
        
        if not all_files:
            self.update_status("Файлы не найдены")
            messagebox.showerror("Ошибка", "Excel-файлы не найдены в указанной папке")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        total_files = len(all_files)
        self.log_message(f"Найдено файлов: {total_files}")
        
        for i, filepath in enumerate(all_files):
            # Обновляем прогресс
            progress_value = (i + 1) / total_files * 100
            self.progress['value'] = progress_value
            
            # Показываем текущий файл
            self.update_status(f"Обработка: {filepath.name}")
            self.log_message(f"[{i+1}/{total_files}] {filepath.name}")
            
            try:
                wb = load_workbook(filepath, data_only=True, read_only=True)
            except Exception as e:
                self.log_message(f"  ✗ Ошибка открытия: {e}")
                continue
            
            # Проверяем наличие листа PLM_TEM_Report
            if "PLM_TEM_Report" not in wb.sheetnames:
                self.log_message(f"  ⊘ Лист PLM_TEM_Report не найден - пропуск")
                wb.close()
                continue
            
            # есть ли в репорте данные
            plm_sheet = wb["PLM_TEM_Report"]
            if plm_sheet.max_row < 6:
                self.log_message(f"  ⊘ Лист PLM_TEM_Report пустой")
                wb.close()
                continue
                
            counter += 1
            # 1. Берем информацию из M1
            project_info = plm_sheet["M1"].value
            date_analyzed = plm_sheet["M2"].value

            
            # 2. Берем информацию из B3 листа SampleAnalyses
            analyst_info = None
            if "SampleAnalyses" in wb.sheetnames:
                sample_sheet = wb["SampleAnalyses"]
                if sample_sheet["B3"].value:
                    if len(str(sample_sheet["B3"].value).strip()) > 1:
                        analyst_info = str(sample_sheet["B3"].value).strip().upper()
                else:
                    analyst_info = "No Analyst"
            else: 
                analyst_info = "No Analyst"
            
            # 3. Считаем строки
            plm_count = 0
            nob_count = 0
            tem_count = 0

            for row in plm_sheet.iter_rows(min_row=6, min_col=9, max_col=17):
                # Колонка I (индекс 0, т.к. min_col=9)
                cell_i = row[0].value
                if cell_i is not None:
                    cell_str = str(cell_i).strip()
                    if cell_str and cell_str != "Not Applicable":
                        plm_count += 1
                
                # Колонка L (индекс 3, т.к. L=12, 12-9=3)
                cell_l = row[3].value
                if cell_l is not None:
                    cell_str = str(cell_l).strip()
                    if cell_str and cell_str != "Not Applicable":
                        nob_count += 1
                
                # Колонка Q (индекс 8, т.к. Q=17, 17-9=8)
                cell_q = row[8].value
                if cell_q is not None:
                    cell_str = str(cell_q).strip()
                    if cell_str and cell_str != "Not Analyzed":
                        tem_count += 1
            
            wb.close()
            
            int_month_analyzed = None
            if date_analyzed:
                # date_analyzed = date(date_analyzed)
                # print(date_analyzed)
                pass
            else:
                date_analyzed = None

            results.append({
                "number": counter,
                "project": project_info,
                "date": date_analyzed, 
                "analyst": analyst_info,
                "plm_count": plm_count,
                "nob_count": nob_count,
                "tem_count": tem_count,
                "month": None
            })
            
            self.log_message(f"  ✓ PLM: {plm_count}, NOB: {nob_count}, TEM: {tem_count}")
        
        if not results:
            self.update_status("Готово (нет данных)")
            messagebox.showwarning("Внимание", "Не найдено файлов с листом PLM_TEM_Report")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        # Сохраняем результат
        self.update_status("Сохраняю файл...")
        self.log_message("\nСохраняю результат...")
        
        output_file = Path(self.folder_path) / "qc_data.xlsx"
        
        result_wb = Workbook()
        sheet_raw = result_wb.active
        sheet_raw.title = "QC Raw Data"
        
        headers = ["number", "project", "date", "analyst", "plm_count", "nob_count", "tem_count", "month"]
        sheet_raw.append(headers)
        
        for r in results:
            sheet_raw.append([
                r["number"],
                r["project"],
                r["date"],
                r["analyst"],
                r["plm_count"],
                r["nob_count"],
                r["tem_count"],
                r["month"]
            ])
        

        result_wb.save(output_file)

        raw_df = pd.read_excel(output_file, sheet_name='QC Raw Data')
        raw_df['date'] = pd.to_datetime(raw_df['date']).dt.date
        raw_df = raw_df.sort_values(by='date')


        ws_raw = result_wb.create_sheet(title='raw_sheet')
        # Заголовки
        ws_raw.append(headers)
        
        for number, row in raw_df.iterrows():
            ws_raw.append([number, row["project"], row["date"],row["analyst"],row["plm_count"],row["nob_count"],row["tem_count"],row["month"],]),

        for analyst in raw_df['analyst'].unique():
            analyst_df = raw_df[raw_df['analyst'] == analyst]
            analyst_df = analyst_df.sort_values(by='date')
            analyst_df = analyst_df.groupby('date', as_index=False)[['plm_count', 'nob_count', 'tem_count']].sum()
            ws_analyst = result_wb.create_sheet(title=f'{analyst}_sum')
            ws_analyst.append(["date", "plm_count", "nob_count", "tem_count"])

            for item, row in analyst_df.iterrows():
                ws_analyst.append([row['date'], row["plm_count"],row["nob_count"],row["tem_count"]])


        result_wb.save(output_file)

        # with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        #     raw_df.to_excel(writer, sheet_name='raw_sheet', index=False)

        #     for analyst in raw_df['analyst'].unique():
        #                 analyst_df = raw_df[raw_df['analyst'] == analyst]
        #                 analyst_df = analyst_df.sort_values(by='date')
        #                 analyst_df = analyst_df.groupby('date', as_index=False)[['plm_count', 'nob_count', 'tem_count']].sum()
        #                 analyst_df.to_excel(writer, sheet_name=f'{analyst}_sum', index=False)


        result_wb = load_workbook(output_file)
        if 'QC Raw Data' in result_wb.sheetnames:
            print('test')
            sheet_to_delete = result_wb['QC Raw Data']
            result_wb.remove(sheet_to_delete)

        
        result_wb.save(output_file)

        # result_wb.remove('QC Raw Data')
        # result_wb.save(output_file)
        # raw_df = pd.DataFrame(results, columns=["number", "project", "date", 
        #                                     "analyst_info", "plm_count", "nob_count", 
        #                                     "tem_count", "month"])
        
        # raw_df['date'] = pd.to_datetime(raw_df['date']).dt.date
        # raw_df = raw_df.sort_values(by='date')

        # with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        #     raw_df.to_excel(writer, sheet_name='raw_sheet', index=False)

        #     for analyst in raw_df['analyst_info'].unique():
        #         analyst_df = raw_df[raw_df['analyst_info'] == analyst]
        #         analyst_df = analyst_df.sort_values(by='date')
        #         analyst_df = analyst_df.groupby('date', as_index=False)[['plm_count', 'nob_count', 'tem_count']].sum()
        #         analyst_df.to_excel(writer, sheet_name=f'{analyst}_sum', index=False)
        
        self.log_message("  ✓ Лист 'По аналитикам' создан")
        self.log_message("  ✓ Лист 'По месяцам' создан")
        self.log_message("  ✓ Лист 'Аналитик-Месяц' создан")
        
        self.progress['value'] = 100
        self.update_status("Готово!")
        self.log_message(f"\n✓ Сохранено: {output_file}")
        self.log_message(f"Обработано файлов с данными: {len(results)}")
        
        messagebox.showinfo(
            "Готово!", 
            f"Обработано файлов: {len(results)}\n\n")
        
        self.btn_run.config(state=NORMAL)
        self.btn_select.config(state=NORMAL)


if __name__ == "__main__":
    root = Tk()
    app = QCProcessorApp(root)
    root.mainloop()

    # pyinstaller --onefile --windowed --name="QC_Collector" qc_info0.07.py