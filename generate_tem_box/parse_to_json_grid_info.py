import openpyxl
import json
from openpyxl import load_workbook
from pathlib import Path
# =============================================== filter blanks
script_dir = Path(__file__).resolve().parent

input_data = None
input_json_path = script_dir / "tem_clean_sorted_data.json"
with open(input_json_path, "r", encoding="utf-8") as f:
    input_data = json.load(f)

tem_grid_data = []
tem_dict = {}


filepath = script_dir / 'grid_id_QC_boxes_TEM.xlsx'
try:
    wb = load_workbook(filepath, data_only=True, read_only=True)
except Exception as e:
    print('ERROR')

ws = wb['QC_boxes']

grid_id_number = None
grid_box_number = None
for block in range(1, ws.max_row, 25):
    grid_box_number = str(ws[f'A{block}'].value)[6:]
    if int(grid_box_number) > 20:
        break
    for row in range(block, (block + 25)):
        if str(ws[f'B{row}'].value) != 'None':
            grid_id_number = ws[f'B{row}'].value
        if str(ws[f'D{row}'].value) != 'None':
            project_with_sample = str(ws[f'D{row}'].value) + '/' + str(ws[f'E{row}'].value)
            
            found_item = next(
                                            (
                                                item
                                                for item in input_data
                                                if item.get("project") == str(ws[f'D{row}'].value)
                                                and item.get("sample") == str(ws[f'E{row}'].value)
                                            ),
                                            None
                                        )
            found_item['Box Number'] = grid_box_number + 'rd'
            found_item['Grid_1'] = ws[f'C{row}'].value + str(grid_id_number)
            if ws[f'C{row}'].value + str(grid_id_number) != 'E5':
                found_item['Grid_2'] = ws[f'C{row+1}'].value + str(grid_id_number)
            else:
                found_item['Grid_2'] = 'A6'
            print(f'found_item : {found_item}')
            tem_grid_data.append(found_item)
    print('=========================================')
    for row in range(block, (block + 25)):
        if str(ws[f'F{row}'].value) != 'None':
            grid_id_number = ws[f'F{row}'].value
        if str(ws[f'H{row}'].value) != 'None':
            if ws[f'G{row}'].value + str(grid_id_number) != 'A6':
                project_with_sample = str(ws[f'H{row}'].value) + '/' + str(ws[f'I{row}'].value)
                found_item = next(
                                                (
                                                    item
                                                    for item in input_data
                                                    if item.get("project") == str(ws[f'H{row}'].value)
                                                    and item.get("sample") == str(ws[f'I{row}'].value)
                                                ),
                                                None
                                            )
                found_item['Box Number'] = grid_box_number + 'rd'
                found_item['Grid_1'] = ws[f'G{row}'].value + str(grid_id_number)
                found_item['Grid_2'] = ws[f'G{row+1}'].value + str(grid_id_number)
                print(f'found_item : {found_item}')
                tem_grid_data.append(found_item)

output_json_path = script_dir / "tem_grid_data.json"
with open(output_json_path, "a", encoding="utf-8") as f:
    f.write(json.dumps(tem_grid_data, indent=4, ensure_ascii=False))