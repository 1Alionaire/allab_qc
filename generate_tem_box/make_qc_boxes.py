# -*- coding: utf-8 -*-
"""
Генерирует страницу QC-боксов из JSON (один массив записей с ключами project/sample).

Использование:
  python make_qc_boxes.py [шаблон.xlsx] [data.json] [выход.xlsx]

Python 3.8+, только openpyxl.
"""
import json
import logging
import math
import sys
import traceback
from copy import copy

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

ROWS_PER_BOX = 25       # один бокс = 25 строк
SAMPLES_PER_BOX = 25    # 10 ячеек x 5 подъячеек / 2 подъячейки на сэмпл
NOTES = [(5, "replicate"), (9, "duplicate"), (13, "blank "), (17, "empty")]


def build_sheet(wb, ws_src, items, sheet_name="QC_boxes"):
    ws = wb.create_sheet(sheet_name)

    # ширины колонок из шаблона
    for letter, dim in ws_src.column_dimensions.items():
        if dim.width is not None:
            ws.column_dimensions[letter].width = dim.width

    n_boxes = max(1, math.ceil(len(items) / SAMPLES_PER_BOX))
    logging.info("%d записей -> %d боксов", len(items), n_boxes)

    for box in range(1, n_boxes + 1):
        # пустая строка-разделитель после каждой пары боксов
        start = (box - 1) * ROWS_PER_BOX + (box - 1) // 2 + 1

        # 1) стили: копируем блок 25x13 из Box #1 шаблона
        for off in range(ROWS_PER_BOX):
            for col in range(1, 14):
                ws.cell(start + off, col)._style = copy(ws_src.cell(1 + off, col)._style)

        # 2) "Box # N" в колонке A
        ws.cell(start, 1).value = "Box # {}".format(box)
        ws.merge_cells(start_row=start, start_column=1,
                       end_row=start + 24, end_column=1)

        # 3) номера ячеек (B: 1-5, F: 6-10), буквы подъячеек (C, G)
        for g in range(5):
            r = start + g * 5
            ws.cell(r, 2).value = g + 1
            ws.cell(r, 6).value = g + 6
            ws.merge_cells(start_row=r, start_column=2, end_row=r + 4, end_column=2)
            ws.merge_cells(start_row=r, start_column=6, end_row=r + 4, end_column=6)
            for col in (3, 7):                      # C и G: A, B, C, D, E
                for k in range(5):
                    ws.cell(r + k, col).value = "ABCDE"[k]

        # 4) пометки replicate / duplicate / blank / empty
        for off, text in NOTES:
            r = start + off
            ws.cell(r, 12).value = text
            ws.merge_cells(start_row=r, start_column=12, end_row=r + 2, end_column=13)

        # 5) сэмплы этого бокса
        chunk = items[(box - 1) * SAMPLES_PER_BOX: box * SAMPLES_PER_BOX]
        for i, item in enumerate(chunk):
            project = str(item["project"])
            sample = str(item["sample"])

            if i <= 11:            # левая половина: D/E, merge 2 строки
                r = start + 2 * i
                ws.cell(r, 4).value = project
                ws.cell(r, 5).value = sample
                ws.merge_cells(start_row=r, start_column=4, end_row=r + 1, end_column=4)
                ws.merge_cells(start_row=r, start_column=5, end_row=r + 1, end_column=5)
            elif i == 12:          # сэмпл на границе: пишем в E5 и в A6
                for r, pc, sc in ((start + 24, 4, 5), (start, 8, 9)):
                    ws.cell(r, pc).value = project
                    ws.cell(r, sc).value = sample
            else:                  # правая половина: H/I, merge 2 строки
                r = start + 1 + 2 * (i - 13)
                ws.cell(r, 8).value = project
                ws.cell(r, 9).value = sample
                ws.merge_cells(start_row=r, start_column=8, end_row=r + 1, end_column=8)
                ws.merge_cells(start_row=r, start_column=9, end_row=r + 1, end_column=9)


def main():
    template_path = sys.argv[1] if len(sys.argv) > 1 else "grid_id_QC_boxes_1-100.xlsx"
    json_path = sys.argv[2] if len(sys.argv) > 2 else "tem_total_data.json"
    out_path = sys.argv[3] if len(sys.argv) > 3 else "grid_id_QC_boxes_TEM.xlsx"

    try:
        with open(json_path, "r", encoding="utf-8") as f:
            items = json.load(f)
        if isinstance(items, dict):            # на случай {"ключ": [...]}
            items = next(iter(items.values()))

        wb = load_workbook(template_path)
        ws_src = wb["Sheet1"]
        build_sheet(wb, ws_src, items)

        # убираем листы шаблона
        for name in ("Sheet1", "Sheet2", "Sheet3"):
            if name in wb.sheetnames:
                del wb[name]

        wb.save(out_path)
        logging.info("Сохранено: %s", out_path)
    except Exception:
        logging.error("Ошибка:\n%s", traceback.format_exc())
        sys.exit(1)


if __name__ == "__main__":
    main()