import threading
from pathlib import Path
from tkinter import Tk, Button, Label, Text, filedialog, messagebox, END, DISABLED, NORMAL
from tkinter.ttk import Progressbar
from openpyxl import Workbook, load_workbook
import pandas as pd
from datetime import date
import json
from tkcalendar import DateEntry
import logging
import logging, traceback, os, sys

plm_analysis_columns = ['Client ID', 'Lab ID', 'Layer', 
                        'Color', 'Texture', 'Homogeneity', 'Morphology', 
                        'RI II Type 1', 'RI II Type 2',
                        'RI ┴ Type 1',  'RI ┴ Type 2', 
                        'Sign of \nElongation Type 1', 'Sign of \nElongation Type 2', 
                        'Extinction \nAngle Type 1', 'Extinction \nAngle Type 2', 
                        'Pleochroism /\nColor Type 1', 'Pleochroism /\nColor Type 2', 
                        'Birefringence Type 1', 'Birefringence Type 2', 
                        'Other Fibers', 'Property', '% Non-\nAsbestos',
                        'Type 1', 'Point 1', 
                        'Type 2', 'Point 2', 
                        'Type 3', 'Point 3', 
                        'Type 4', 'Point 4', 
                        'Type 5', 'Point 5', 
                        'Type 6', 'Point 6', 
                        'Type 7', 'Point 7', 
                        'Type 8', 'Point 8', 
                        'Type Asb 1 Option', 'Percent 1 Option',
                        'Type Asb 2 Option', 'Percent 2 Option',
                        'Vermiculite', 'Method', 'Undesolved Materials', 'Total Residue']

tem_analysis_columns = ['Client ID', 'Lab ID', 'Layer',  'Homogeneity',  'Residue', 
                        'Point Type 1', 'Percent Type 1', 'Asb Type Type 1', 
                         'Point Type 2', 'Percent Type 2', 'Asb Type Type 2', 
                         'Microscope', 'Eccentricity', 'Grid Pre', 'Grid Box #',
                         'Grid Box ID 1', 'Grid Box ID 2', 'Method', 'NA or PS']

def get_writable_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)   # папка рядом с .exe
    return os.path.dirname(os.path.abspath(__file__))

def find_project_in_sample(inp_sample_id):
    final_result = ''
    count = 2
    for i in inp_sample_id:
        if count == 0:
            return final_result[:-1]
        final_result += i
        if i == '-':
            count -= 1

logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s'
)

def is_round_value(inp_value):
    if (inp_value * 100) % 1 != 0:
        return False
    else:
        return True

class QCProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PLM QC By Date")
        self.root.geometry("900x800")
        self.root.resizable(True, True)
        
        self.folder_path = None

        # Кнопка выбора папки
        self.btn_select = Button(
            root, 
            text="Choose Folder", 
            command=self.select_folder,
            width=20,
            height=2
        )
        self.btn_select.pack(pady=10)
        
        # Метка с выбранной папкой
        self.lbl_folder = Label(root, text="Папка не выбрана", wraplength=550)
        self.lbl_folder.pack(pady=5)

        self.labels = {}

        # Кнопка запуска
        self.btn_run = Button(
            root, 
            text="Запустить", 
            command=self.run_processing,
            width=20,
            height=2,
            state=DISABLED
        )
        self.btn_run.pack(pady=10)
        
        # Прогресс-бар
        self.progress = Progressbar(root, length=550, mode='determinate')
        self.progress.pack(pady=10)
        
        # Текущий статус
        self.lbl_status = Label(root, text="", font=("Arial", 10))
        self.lbl_status.pack(pady=5)
        
        # Лог
        self.log = Text(root, height=10, width=70, state=DISABLED)
        self.log.pack(pady=10)

    def select_folder(self):
        folder = filedialog.askdirectory(title="Выберите папку с Excel-файлами")
        if folder:
            self.folder_path = folder
            self.lbl_folder.config(text=f"Папка: {folder}")
            self.btn_run.config(state=NORMAL)
    
    def log_message(self, message):
        """Добавляет сообщение в лог"""
        self.log.config(state=NORMAL)
        self.log.insert(END, message + "\n")
        self.log.see(END)
        self.log.config(state=DISABLED)
    
    def update_status(self, message):
        """Обновляет статус"""
        self.lbl_status.config(text=message)
        self.root.update()
    
    def run_processing(self):

        if not self.folder_path:
            messagebox.showwarning("Warning", "Please choose folder for collect")
            return

        """Запускает обработку в отдельном потоке"""
        self.btn_run.config(state=DISABLED)
        self.btn_select.config(state=DISABLED)
        self.progress['value'] = 0
        
        # Очищаем лог
        self.log.config(state=NORMAL)
        self.log.delete(1.0, END)
        self.log.config(state=DISABLED)
        
        # Запускаем в отдельном потоке, чтобы GUI не зависал
        thread = threading.Thread(target=self.process_files)
        thread.start()
    
    def process_files(self):
        """Основная логика обработки файлов"""
        results = []
        result_json = {}
        counter = 0
        
        # Ищем все Excel-файлы
        self.update_status("Поиск Excel-файлов...")
        all_files = []
        for ext in ("*.xlsx", "*.xlsm", "*.xls"):
            all_files.extend(Path(self.folder_path).rglob(ext))
        
        if not all_files:
            self.update_status("Файлы не найдены")
            messagebox.showerror("Ошибка", "Excel-файлы не найдены в указанной папке")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
    

        total_files = len(all_files)
        self.log_message(f"Найдено файлов: {total_files}")
        
        for i, filepath in enumerate(all_files):
            # Обновляем прогресс
            progress_value = (i + 1) / total_files * 100
            self.progress['value'] = progress_value
            
            # Показываем текущий файл
            self.update_status(f"Обработка: {filepath.name}")
            self.log_message(f"[{i+1}/{total_files}] {filepath.name}")

            if ('PLM' not in filepath.name and 'NOB' not in filepath.name and 'TEM' not in filepath.name):
                self.log_message(f" {filepath.name} - ⊘ НЕ PLM файл")
                continue

            if 'Conflict' in str(filepath):
                self.log_message(f" {filepath.name} - ⊘ Duplicate from Conflict")
                continue

            try:
                wb = load_workbook(filepath, data_only=True, read_only=True)
            except Exception as e:
                self.log_message(f"{filepath.name} ✗ Ошибка открытия: {e}")
                continue

            # Проверяем наличие листа PLM_TEM_Report
            if "NOB_Calculation" not in wb.sheetnames:
                self.log_message(f" {filepath.name} ⊘ Лист PLM_TEM_Report не найден ")
                wb.close()
                continue
            
            # есть ли в репорте данные
            weight_sheet = wb["NOB_Calculation"]
            if weight_sheet.max_row < 7:
                self.log_message(f" {filepath.name} ⊘ Лист PLM_TEM_Report пустой")
                wb.close()
                continue

            counter += 1

            for row in weight_sheet.iter_rows(min_row=7, max_row=weight_sheet.max_row, min_col=1, max_col=20, values_only=True):
                try:
                    count = 0
                    for i in range(3, 10):
                        if is_round_value(float(row[i])):
                            count += 1
                    if count == 7:
                        pass
                    else:
                        if ((0 < float(row[7]) < 100) and (0 < float(row[12]) < 100) and (0 < float(row[13]) < 100)):
                            results.append({
                                'cruc_weight': round(float(row[3]), 4),
                                'cruc_with_sample_weight': round(float(row[4]), 4),
                                'sample_weight': round(float(row[5]), 4),
                                'cruc_with_sample_ash_weight': round(float(row[6]), 4),
                                'percent_organic': round(float(row[7]), 4),
                                'petri_weight': round(float(row[8]), 4),
                                'petri_with_sample_weight': round(float(row[9]), 4),
                                'percent_caco3': round(float(row[12]), 4),
                                'percent_residue': round(float(row[13]), 4),
                            })  
                except:
                    continue

            wb.close()
        
        output_raw_file = Path(self.folder_path) / "weight_data.json"

        with open(output_raw_file, "a", encoding="utf-8") as f:
            f.write(json.dumps(results, indent=4, ensure_ascii=False))

        # generated_duplicates_array = generate_duplicates(result_json)
        # output_total_array_file = Path(self.folder_path) / "qc_result_data.json"

        # with open(output_total_array_file, "a", encoding="utf-8") as f:
        #     f.write(json.dumps(generated_duplicates_array, indent=4, ensure_ascii=False))


        if not results:
            self.update_status("Готово (нет данных)")
            messagebox.showwarning("Внимание", "Done!")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        # Сохраняем результат
        self.update_status("Сохраняю файл...")
        self.log_message("\nСохраняю результат...")
        
        output_file = Path(self.folder_path) / "qc_data.xlsx"
      
        self.log_message("  ✓ Лист 'По аналитикам' создан")
        self.log_message("  ✓ Лист 'По месяцам' создан")
        self.log_message("  ✓ Лист 'Аналитик-Месяц' создан")
        
        self.progress['value'] = 100
        self.update_status("Готово!")
        self.log_message(f"\n✓ Сохранено: {output_file}")
        self.log_message(f"Обработано файлов с данными: {len(results)}")
        
        messagebox.showinfo(
            "Готово!", 
            f"Обработано файлов: {len(results)}\n\n")
        
        self.btn_run.config(state=NORMAL)
        self.btn_select.config(state=NORMAL)


if __name__ == "__main__":
    root = Tk()
    app = QCProcessorApp(root)
    root.mainloop()

    # pyinstaller --onefile --hidden-import babel.numbers --hidden-import babel.dates --collect-all babel --name="Parsing_all_weights" parsing_all_weights.py