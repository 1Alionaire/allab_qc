import threading
from pathlib import Path
from tkinter import Tk, Button, Label, Text, filedialog, messagebox, END, DISABLED, NORMAL
from tkinter.ttk import Progressbar
from openpyxl import Workbook, load_workbook
import pandas as pd
from datetime import date
import json
from tkcalendar import DateEntry
from datetime import datetime
import logging
import logging, traceback, os, sys

plm_analysis_columns = ['Client ID', 'Lab ID', 'Layer', 
                        'Color', 'Texture', 'Homogeneity', 'Morphology', 
                        'RI II Type 1', 'RI II Type 2',
                        'RI ┴ Type 1',  'RI ┴ Type 2', 
                        'Sign of \nElongation Type 1', 'Sign of \nElongation Type 2', 
                        'Extinction \nAngle Type 1', 'Extinction \nAngle Type 2', 
                        'Pleochroism /\nColor Type 1', 'Pleochroism /\nColor Type 2', 
                        'Birefringence Type 1', 'Birefringence Type 2', 
                        'Other Fibers', 'Property', '% Non-\nAsbestos',
                        'Type 1', 'Point 1', 
                        'Type 2', 'Point 2', 
                        'Type 3', 'Point 3', 
                        'Type 4', 'Point 4', 
                        'Type 5', 'Point 5', 
                        'Type 6', 'Point 6', 
                        'Type 7', 'Point 7', 
                        'Type 8', 'Point 8', 
                        'Type Asb 1 Option', 'Percent 1 Option',
                        'Type Asb 2 Option', 'Percent 2 Option',
                        'Vermiculite', 'Method', 'Undesolved Materials', 'Total Residue']

tem_analysis_columns = ['Client ID', 'Lab ID', 'Layer',  'Homogeneity',  'Residue', 
                        'Point Type 1', 'Percent Type 1', 'Asb Type Type 1', 
                         'Point Type 2', 'Percent Type 2', 'Asb Type Type 2', 
                         'Microscope', 'Eccentricity', 'Grid Pre', 'Grid Box #',
                         'Grid Box ID 1', 'Grid Box ID 2', 'Method', 'NA or PS']

def get_writable_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)   # папка рядом с .exe
    return os.path.dirname(os.path.abspath(__file__))

def find_project_in_sample(inp_sample_id):
    final_result = ''
    count = 2
    for i in inp_sample_id:
        if count == 0:
            return final_result[:-1]
        final_result += i
        if i == '-':
            count -= 1

logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s'
)


class QCProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PLM QC By Date")
        self.root.geometry("900x800")
        self.root.resizable(True, True)
        
        self.folder_path = None

        self.select_date = DateEntry(self.root, width=18, background='darkblue',
                                       foreground='white', borderwidth=2,
                                       date_pattern='mm.dd.yyyy')
        self.select_date.pack(pady=5)

        # Кнопка выбора папки
        self.btn_select = Button(
            root, 
            text="Choose Folder", 
            command=self.select_folder,
            width=20,
            height=2
        )
        self.btn_select.pack(pady=10)
        
        # Метка с выбранной папкой
        self.lbl_folder = Label(root, text="Папка не выбрана", wraplength=550)
        self.lbl_folder.pack(pady=5)

        self.labels = {}
        for key, caption in [("plm", "PLM File"),
                             ("nob", "NOB File"),
                             ("tem", "TEM File")]:
            Button(root, text=f"Choose {caption}", width=20, height=2,
                   command=lambda k=key: self.select_file(k)).pack(pady=5)
            lbl = Label(root, text=f"{caption}: не выбран", wraplength=550)
            lbl.pack(pady=2)
            self.labels[key] = lbl
        
        # Кнопка запуска
        self.btn_run = Button(
            root, 
            text="Запустить", 
            command=self.run_processing,
            width=20,
            height=2,
            state=DISABLED
        )
        self.btn_run.pack(pady=10)
        
        # Прогресс-бар
        self.progress = Progressbar(root, length=550, mode='determinate')
        self.progress.pack(pady=10)
        
        # Текущий статус
        self.lbl_status = Label(root, text="", font=("Arial", 10))
        self.lbl_status.pack(pady=5)
        
        # Лог
        self.log = Text(root, height=10, width=70, state=DISABLED)
        self.log.pack(pady=10)

    def select_folder(self):
        folder = filedialog.askdirectory(title="Выберите папку с Excel-файлами")
        if folder:
            self.folder_path = folder
            self.lbl_folder.config(text=f"Папка: {folder}")
            self.btn_run.config(state=NORMAL)
    
    def log_message(self, message):
        """Добавляет сообщение в лог"""
        self.log.config(state=NORMAL)
        self.log.insert(END, message + "\n")
        self.log.see(END)
        self.log.config(state=DISABLED)
    
    def update_status(self, message):
        """Обновляет статус"""
        self.lbl_status.config(text=message)
        self.root.update()
    
    def run_processing(self):

        if not self.folder_path:
            messagebox.showwarning("Warning", "Please choose folder for collect")
            return

        """Запускает обработку в отдельном потоке"""
        self.btn_run.config(state=DISABLED)
        self.btn_select.config(state=DISABLED)
        self.progress['value'] = 0
        
        # Очищаем лог
        self.log.config(state=NORMAL)
        self.log.delete(1.0, END)
        self.log.config(state=DISABLED)
        
        # Запускаем в отдельном потоке, чтобы GUI не зависал
        thread = threading.Thread(target=self.process_files)
        thread.start()
    
    def process_files(self):
        """Основная логика обработки файлов"""
        results = []
        result_json = {}
        counter = 0
        wrong_files_array = []
        
        # Ищем все Excel-файлы
        self.update_status("Поиск Excel-файлов...")
        all_files = []
        for ext in ("*.xlsx", "*.xlsm", "*.xls"):
            all_files.extend(Path(self.folder_path).rglob(ext))
        
        if not all_files:
            self.update_status("Файлы не найдены")
            messagebox.showerror("Ошибка", "Excel-файлы не найдены в указанной папке")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
    

        total_files = len(all_files)
        self.log_message(f"Найдено файлов: {total_files}")
        
        for i, filepath in enumerate(all_files):
            # Обновляем прогресс
            progress_value = (i + 1) / total_files * 100
            self.progress['value'] = progress_value
            
            # Показываем текущий файл
            self.update_status(f"Обработка: {filepath.name}")
            self.log_message(f"[{i+1}/{total_files}] {filepath.name}")

            if ('NOB' not in filepath.name and 'TEM' not in filepath.name):
                self.log_message(f" {filepath.name} - ⊘ НЕ PLM файл")
                continue

            if 'Conflict' in str(filepath):
                self.log_message(f" {filepath.name} - ⊘ Duplicate from Conflict")
                continue

            try:
                wb = load_workbook(filepath, data_only=True, read_only=True)
            except Exception as e:
                self.log_message(f"{filepath.name} ✗ Ошибка открытия: {e}")
                continue

            # Проверяем наличие листа PLM_TEM_Report
            if "PLM_TEM_Report" not in wb.sheetnames:
                self.log_message(f" {filepath.name} ⊘ Лист PLM_TEM_Report не найден ")
                wb.close()
                continue

            # есть ли в репорте данные
            plm_result_sheet = wb["PLM_TEM_Report"]
            if plm_result_sheet.max_row < 6:
                self.log_message(f" {filepath.name} ⊘ Лист PLM_TEM_Report пустой")
                wb.close()
                continue
            
            counter += 1
            project_info = plm_result_sheet["M1"].value

            if project_info is None or str(project_info).strip() == '' or str(project_info).strip() == 'None':
                project_info = find_project_in_sample(str(plm_result_sheet["B6"].value))

            if project_info in result_json:
                self.log_message(f" {filepath.name} ⊘ Already has record")
                wb.close()
                continue
            
            total_count = 0
            # 2. Берем информацию из B3 листа SampleAnalyses
            analyst_info = None
            for row in plm_result_sheet.iter_rows(min_row=6, min_col=2, max_col=2):
                cell_i = row[0].value
                if cell_i is not None:
                    cell_str = str(cell_i).strip()
                    if (cell_str):
                        if (cell_str) and (project_info in cell_str):
                            total_count += 1

            all_tem_data = []

            if "TEM_Calculation" in wb.sheetnames:
                tem_sample_sheet = wb["TEM_Calculation"]

                if tem_sample_sheet['A6'].value is not None:
                    for row in tem_sample_sheet.iter_rows(min_row=6, max_row=(6 + total_count - 1), min_col=1, max_col=19, values_only=True):
                        tem_analysis_json = {}
                        if row[0] is None:
                            continue
                        elif row[18] == 'PS':
                            continue
                        elif str(row[0]).lower() == 'bl':
                            if str(row[15]).lower() != 'none' and str(row[16]).lower() != 'none':

                                result_json[project_info] = {'grid_box': str(row[14]).lower(),
                                                            'grid_id_1:' : str(row[15]).lower(),
                                                            'grid_id_2:' : str(row[16]).lower()}
            wb.close()

        output_total_array_file = Path(self.folder_path) / "qc_result_data.json"
        with open(output_total_array_file, "a", encoding="utf-8") as f:
            f.write(json.dumps(result_json, indent=4, ensure_ascii=False))
        
        with open("wrong data files.txt", "w", encoding="utf-8") as file:
            file.writelines(f"{item}\n" for item in wrong_files_array)

        if not results:
            self.update_status("Готово (нет данных)")
            messagebox.showwarning("Внимание", "Done!")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        # Сохраняем результат
        self.update_status("Сохраняю файл...")
        self.log_message("\nСохраняю результат...")
        
        output_file = Path(self.folder_path) / "qc_data.xlsx"
      
        self.log_message("  ✓ Лист 'По аналитикам' создан")
        self.log_message("  ✓ Лист 'По месяцам' создан")
        self.log_message("  ✓ Лист 'Аналитик-Месяц' создан")
        
        self.progress['value'] = 100
        self.update_status("Готово!")
        self.log_message(f"\n✓ Сохранено: {output_file}")
        self.log_message(f"Обработано файлов с данными: {len(results)}")
        
        messagebox.showinfo(
            "Готово!", 
            f"Обработано файлов: {len(results)}\n\n")
        
        self.btn_run.config(state=NORMAL)
        self.btn_select.config(state=NORMAL)


if __name__ == "__main__":
    root = Tk()
    app = QCProcessorApp(root)
    root.mainloop()

    # pyinstaller --onefile --hidden-import babel.numbers --hidden-import babel.dates --collect-all babel --add-data "correct_weight.json;." --name="PLM_QC_Generate_by_day" main.py