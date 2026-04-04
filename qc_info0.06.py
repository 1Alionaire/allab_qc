import threading
from pathlib import Path
from tkinter import Tk, Button, Label, Text, filedialog, messagebox, END, DISABLED, NORMAL
from tkinter.ttk import Progressbar
from openpyxl import Workbook, load_workbook
import pandas as pd
from datetime import date


class QCProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("QC Processor")
        self.root.geometry("600x400")
        self.root.resizable(False, False)
        
        self.folder_path = None
        
        # Кнопка выбора папки
        self.btn_select = Button(
            root, 
            text="Choose Folder", 
            command=self.select_folder,
            width=20,
            height=2
        )
        self.btn_select.pack(pady=10)
        
        # Метка с выбранной папкой
        self.lbl_folder = Label(root, text="Папка не выбрана", wraplength=550)
        self.lbl_folder.pack(pady=5)
        
        # Кнопка запуска
        self.btn_run = Button(
            root, 
            text="Запустить", 
            command=self.run_processing,
            width=20,
            height=2,
            state=DISABLED
        )
        self.btn_run.pack(pady=10)
        
        # Прогресс-бар
        self.progress = Progressbar(root, length=550, mode='determinate')
        self.progress.pack(pady=10)
        
        # Текущий статус
        self.lbl_status = Label(root, text="", font=("Arial", 10))
        self.lbl_status.pack(pady=5)
        
        # Лог
        self.log = Text(root, height=10, width=70, state=DISABLED)
        self.log.pack(pady=10)
    
    def select_folder(self):
        folder = filedialog.askdirectory(title="Выберите папку с Excel-файлами")
        if folder:
            self.folder_path = folder
            self.lbl_folder.config(text=f"Папка: {folder}")
            self.btn_run.config(state=NORMAL)
    
    def log_message(self, message):
        """Добавляет сообщение в лог"""
        self.log.config(state=NORMAL)
        self.log.insert(END, message + "\n")
        self.log.see(END)
        self.log.config(state=DISABLED)
    
    def update_status(self, message):
        """Обновляет статус"""
        self.lbl_status.config(text=message)
        self.root.update()
    
    def run_processing(self):
        """Запускает обработку в отдельном потоке"""
        self.btn_run.config(state=DISABLED)
        self.btn_select.config(state=DISABLED)
        self.progress['value'] = 0
        
        # Очищаем лог
        self.log.config(state=NORMAL)
        self.log.delete(1.0, END)
        self.log.config(state=DISABLED)
        
        # Запускаем в отдельном потоке, чтобы GUI не зависал
        thread = threading.Thread(target=self.process_files)
        thread.start()
    
    def process_files(self):
        """Основная логика обработки файлов"""
        results = []
        counter = 0
        
        # Ищем все Excel-файлы
        self.update_status("Поиск Excel-файлов...")
        all_files = []
        for ext in ("*.xlsx", "*.xlsm", "*.xls"):
            all_files.extend(Path(self.folder_path).rglob(ext))
        
        if not all_files:
            self.update_status("Файлы не найдены")
            messagebox.showerror("Ошибка", "Excel-файлы не найдены в указанной папке")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        total_files = len(all_files)
        self.log_message(f"Найдено файлов: {total_files}")
        
        for i, filepath in enumerate(all_files):
            # Обновляем прогресс
            progress_value = (i + 1) / total_files * 100
            self.progress['value'] = progress_value
            
            # Показываем текущий файл
            self.update_status(f"Обработка: {filepath.name}")
            self.log_message(f"[{i+1}/{total_files}] {filepath.name}")
            
            try:
                wb = load_workbook(filepath, data_only=True, read_only=True)
            except Exception as e:
                self.log_message(f"  ✗ Ошибка открытия: {e}")
                continue
            
            # Проверяем наличие листа PLM_TEM_Report
            if "PLM_TEM_Report" not in wb.sheetnames:
                self.log_message(f"  ⊘ Лист PLM_TEM_Report не найден - пропуск")
                wb.close()
                continue
            
            # есть ли в репорте данные
            plm_sheet = wb["PLM_TEM_Report"]
            if plm_sheet.max_row < 6:
                self.log_message(f"  ⊘ Лист PLM_TEM_Report пустой")
                wb.close()
                continue
                
            counter += 1
            # 1. Берем информацию из M1
            project_info = plm_sheet["M1"].value
            date_analyzed = plm_sheet["M2"].value

            
            # 2. Берем информацию из B3 листа SampleAnalyses
            analyst_info = None
            if "SampleAnalyses" in wb.sheetnames:
                sample_sheet = wb["SampleAnalyses"]
                if sample_sheet["B3"].value:
                    if len(str(sample_sheet["B3"].value).strip()) > 1:
                        analyst_info = str(sample_sheet["B3"].value).strip()
                else:
                    analyst_info = "No Analyst"
            else: 
                analyst_info = "No Analyst"
            
            # 3. Считаем строки
            plm_count = 0
            nob_count = 0
            tem_count = 0

            for row in plm_sheet.iter_rows(min_row=6, min_col=9, max_col=17):
                # Колонка I (индекс 0, т.к. min_col=9)
                cell_i = row[0].value
                if cell_i is not None:
                    cell_str = str(cell_i).strip()
                    if cell_str and cell_str != "Not Applicable":
                        plm_count += 1
                
                # Колонка L (индекс 3, т.к. L=12, 12-9=3)
                cell_l = row[3].value
                if cell_l is not None:
                    cell_str = str(cell_l).strip()
                    if cell_str and cell_str != "Not Applicable":
                        nob_count += 1
                
                # Колонка Q (индекс 8, т.к. Q=17, 17-9=8)
                cell_q = row[8].value
                if cell_q is not None:
                    cell_str = str(cell_q).strip()
                    if cell_str and cell_str != "Not Analyzed":
                        tem_count += 1
            
            wb.close()
            
            int_month_analyzed = None
            if date_analyzed:
                # date_analyzed = date(date_analyzed)
                print(date_analyzed)
            else:
                date_analyzed = None

            results.append({
                "number": counter,
                "project": project_info,
                "date": date_analyzed, 
                "analyst_info": analyst_info,
                "plm_count": plm_count,
                "nob_count": nob_count,
                "tem_count": tem_count,
                "month": None
            })
            
            self.log_message(f"  ✓ PLM: {plm_count}, NOB: {nob_count}, TEM: {tem_count}")
        
        if not results:
            self.update_status("Готово (нет данных)")
            messagebox.showwarning("Внимание", "Не найдено файлов с листом PLM_TEM_Report")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        # Сохраняем результат
        self.update_status("Сохраняю файл...")
        self.log_message("\nСохраняю результат...")
        
        output_file = Path(self.folder_path) / "qc_data.xlsx"
        
        out_wb = Workbook()
        out_sheet = out_wb.active
        out_sheet.title = "QC Data"
        
        headers = ["Number", "Project", "Date", "Analyst", "plm_count", "nob_count", "tem_count", "Month"]
        out_sheet.append(headers)
        
        for r in results:
            out_sheet.append([
                r["number"],
                r["project"],
                r["date"],
                r["analyst_info"],
                r["plm_count"],
                r["nob_count"],
                r["tem_count"],
                r["month"]
            ])
        
        for col in out_sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            out_sheet.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        out_wb.save(output_file)
        
        # === ЭТАП 3: Группировка с помощью pandas ===
        self.update_status("Создаю сводные таблицы...")
        self.log_message("\nСоздаю сводные таблицы...")
        
        # Читаем созданный файл
        df = pd.read_excel(output_file)
        
        # Очищаем данные: убираем пробелы и приводим к единому формату
        df["Analyst"] = df["Analyst"].astype(str).str.strip().str.upper()
        df["Month"] = df["Month"].astype(str).str.strip()
        
        # Группировка по Analyst (колонка C = Analyst)
        analyst_summary = df.groupby("Analyst").agg({
            "plm_count": "sum",
            "nob_count": "sum",
            "tem_count": "sum",
            "Number": "count"
        }).rename(columns={"Number": "Кол-во проектов"}).reset_index()
        analyst_summary.rename(columns={"Analyst": "Analyst"}, inplace=True)
        
        # Группировка по Month (колонка G = Month)
        month_summary = df.groupby("Month").agg({
            "plm_count": "sum",
            "nob_count": "sum",
            "tem_count": "sum",
            "Number": "count"
        }).rename(columns={"Number": "Кол-во проектов"}).reset_index()
        
        # Группировка по Analyst + Month (разбивка по месяцам у каждого аналитика)
        analyst_month_summary = df.groupby(["Analyst", "Month"]).agg({
            "plm_count": "sum",
            "nob_count": "sum",
            "tem_count": "sum",
            "Number": "count"
        }).rename(columns={"Number": "Кол-во проектов"}).reset_index()
        analyst_month_summary.rename(columns={"Analyst": "Analyst"}, inplace=True)
        
        # Записываем сводные таблицы в тот же файл на отдельные листы
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            analyst_summary.to_excel(writer, sheet_name='По аналитикам', index=False)
            month_summary.to_excel(writer, sheet_name='По месяцам', index=False)
            analyst_month_summary.to_excel(writer, sheet_name='Аналитик-Месяц', index=False)
        
        self.log_message("  ✓ Лист 'По аналитикам' создан")
        self.log_message("  ✓ Лист 'По месяцам' создан")
        self.log_message("  ✓ Лист 'Аналитик-Месяц' создан")
        
        self.progress['value'] = 100
        self.update_status("Готово!")
        self.log_message(f"\n✓ Сохранено: {output_file}")
        self.log_message(f"Обработано файлов с данными: {len(results)}")
        
        messagebox.showinfo(
            "Готово!", 
            f"Обработано файлов: {len(results)}\n\n"
            f"Созданы листы:\n"
            f"• QC Data - все данные\n"
            f"• По аналитикам - группировка по Analyst\n"
            f"• По месяцам - группировка по Month\n"
            f"• Аналитик-Месяц - разбивка по месяцам у каждого аналитика\n\n"
            f"Файл: {output_file}"
        )
        
        self.btn_run.config(state=NORMAL)
        self.btn_select.config(state=NORMAL)


if __name__ == "__main__":
    root = Tk()
    app = QCProcessorApp(root)
    root.mainloop()