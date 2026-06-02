import json
from pathlib import Path
import math
import random
import copy
import os
import logging
import sys
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
from utils.utilities import calc_std_dev


logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s'
)

def main_start():
    script_dir = Path(__file__).resolve().parent
    resource_dir = script_dir.parent / "resource"

    # template_excel_path = Path(resource_dir / 'Template.xlsx')
    # wb = load_workbook(template_excel_path)

    # Выбираем лист
    

    for file in script_dir.glob("*.json"):
        file_base_name = os.path.basename(file)

        template_excel_path = Path(resource_dir / 'Template.xlsx')
        wb = load_workbook(template_excel_path)

        with file.open("r", encoding="utf-8") as f:
            duplicates_data = json.load(f)
        
        for analyst, info in duplicates_data.items():
            if analyst == 'total':
                ws = wb["Reps"]
                count = 4
                row_number = 1 
                for item_info in info:
                    text_date_for_excel = item_info['date'][5:7] + '-' + item_info['date'][8:10] + '-' + item_info['date'][0:4]
                    ws[f'A{count}'] = row_number
                    ws[f'B{count}'] = text_date_for_excel
                    ws[f'C{count}'] = item_info['project']
                    ws[f'D{count}'] = item_info['sample']
                    ws[f'E{count}'] = item_info['1st_analyst_name']
                    ws[f'F{count}'] = item_info['1st_analyst_asb_type'][:4] 
                    ws[f'G{count}'] = item_info['1st_analyst_asb_percent']
                    ws[f'H{count}'] = text_date_for_excel #
                    ws[f'I{count}'] = item_info['dup_name'] #
                    ws[f'J{count}'] = item_info['dup_asb_type'][:4] #
                    ws[f'K{count}'] = item_info['dup_asb_percent'] #
                    ws[f'L{count}'] = calc_std_dev(item_info['1st_analyst_asb_percent'], item_info['dup_asb_percent'])
                    count += 1
                    row_number += 1

        wb.save('plm_rep' + file_base_name.replace('json', 'xlsx'))
        
if __name__ == "__main__":
    main_start()