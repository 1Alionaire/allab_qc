import os
from pathlib import Path
from openpyxl import Workbook, load_workbook

def process_excel_files(root_folder, output_file="qc_data.xlsx"):
    results = []
    counter = 0
    
    for filepath in Path(root_folder).rglob("*.xlsx"):
        try:
            wb = load_workbook(filepath, data_only=True, read_only=True)
        except Exception as e:
            print(f"Не удалось открыть {filepath}: {e}")
            continue
        
        # Проверяем наличие листа PLM_TEM_Report
        if "PLM_TEM_Report" not in wb.sheetnames:
            wb.close()
            continue
        
        counter += 1
        
        # 1. Берем информацию из M1 листа PLM_TEM_Report
        plm_sheet = wb["PLM_TEM_Report"]
        project_info = plm_sheet["M1"].value
        
        # 2. Берем информацию из B3 листа SampleAnalyses
        sample_info = None
        if "SampleAnalyses" in wb.sheetnames:
            sample_sheet = wb["SampleAnalyses"]
            sample_info = sample_sheet["B3"].value
        
        # 3. Считаем строки по значениям в колонке G (начиная с 6 строки)
        plm_count = 0
        nob_count = 0
        tem_count = 0
        
        for row in plm_sheet.iter_rows(min_row=6, min_col=7, max_col=7):
            cell_value = row[0].value
            if cell_value is not None:
                cell_str = str(cell_value).strip()
                if cell_str in ("198.1", "198,1"):
                    plm_count += 1
                elif cell_str in ("198.6", "198,6"):
                    nob_count += 1
                elif cell_str in ("198.4", "198,4"):
                    tem_count += 1
        
        wb.close()
        
        results.append({
            "number": counter,
            "project": project_info,
            "sample_info": sample_info,
            "plm_count": plm_count,
            "nob_count": nob_count,
            "tem_count": tem_count
        })
        
        print(f"Обработан: {filepath.name}")
    
    # Создаем выходной файл
    out_wb = Workbook()
    out_sheet = out_wb.active
    out_sheet.title = "QC Data"
    
    # Заголовки
    headers = ["Номер", "Проект", "SampleAnalyses B3", "plm_count", "nob_count", "tem_count"]
    out_sheet.append(headers)
    
    # Данные
    for r in results:
        out_sheet.append([
            r["number"],
            r["project"],
            r["sample_info"],
            r["plm_count"],
            r["nob_count"],
            r["tem_count"]
        ])
    
    # Автоширина колонок
    for col in out_sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        out_sheet.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    out_wb.save(output_file)
    print(f"\nГотово! Обработано файлов: {len(results)}")
    print(f"Результат сохранен в: {output_file}")

if __name__ == "__main__":
    # Укажите путь к корневой папке с Excel-файлами
    ROOT_FOLDER = "/path/to/your/excel/files"
    
    process_excel_files(ROOT_FOLDER, "qc_data.xlsx")