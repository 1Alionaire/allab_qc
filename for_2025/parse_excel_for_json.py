import threading
from pathlib import Path
from tkinter import Tk, Button, Label, Text, filedialog, messagebox, END, DISABLED, NORMAL
from tkinter.ttk import Progressbar
from openpyxl import Workbook, load_workbook
import pandas as pd
from datetime import date
import json

plm_analysis_columns = ['Client ID', 'Lab ID', 'Layer', 
                        'Color', 'Texture', 'Homogeneity', 'Morphology', 
                        'RI II Type 1', 'RI II Type 2',
                        'RI ┴ Type 1',  'RI ┴ Type 2', 
                        'Sign of \nElongation Type 1', 'Sign of \nElongation Type 2', 
                        'Extinction \nAngle Type 1', 'Extinction \nAngle Type 2', 
                        'Pleochroism /\nColor Type 1', 'Pleochroism /\nColor Type 2', 
                        'Birefringence Type 1', 'Birefringence Type 2', 
                        'Other Fibers', 'Property', '% Non-\nAsbestos',
                        'Type 1', 'Point 1', 
                        'Type 2', 'Point 2', 
                        'Type 3', 'Point 3', 
                        'Type 4', 'Point 4', 
                        'Type 5', 'Point 5', 
                        'Type 6', 'Point 6', 
                        'Type 7', 'Point 7', 
                        'Type 8', 'Point 8', 
                        'Type Asb 1 Option', 'Percent 1 Option',
                        'Type Asb 2 Option', 'Percent 2 Option',
                        'Vermiculite', 'Method', 'Undesolved Materials', 'Total Residue']

tem_analysis_columns = ['Client ID', 'Lab ID', 'Layer',  'Homogeneity',  'Residue', 
                        'Point Type 1', 'Percent Type 1', 'Asb Type Type 1', 
                         'Point Type 2', 'Percent Type 2', 'Asb Type Type 2', 
                         'Microscope', 'Eccentricity', 'Grid Pre', 'Grid Box #',
                         'Grid Box ID 1', 'Grid Box ID 2', 'Method', 'NA or PS']

def find_project_in_sample(inp_sample_id):
    final_result = ''
    count = 2
    for i in inp_sample_id:
        if count == 0:
            return final_result[:-1]
        final_result += i
        if i == '-':
            count -= 1

class QCProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("QC Processor")
        self.root.geometry("600x400")
        self.root.resizable(False, False)
        
        self.folder_path = None
        
        # Кнопка выбора папки
        self.btn_select = Button(
            root, 
            text="Choose Folder", 
            command=self.select_folder,
            width=20,
            height=2
        )
        self.btn_select.pack(pady=10)
        
        # Метка с выбранной папкой
        self.lbl_folder = Label(root, text="Папка не выбрана", wraplength=550)
        self.lbl_folder.pack(pady=5)
        
        # Кнопка запуска
        self.btn_run = Button(
            root, 
            text="Запустить", 
            command=self.run_processing,
            width=20,
            height=2,
            state=DISABLED
        )
        self.btn_run.pack(pady=10)
        
        # Прогресс-бар
        self.progress = Progressbar(root, length=550, mode='determinate')
        self.progress.pack(pady=10)
        
        # Текущий статус
        self.lbl_status = Label(root, text="", font=("Arial", 10))
        self.lbl_status.pack(pady=5)
        
        # Лог
        self.log = Text(root, height=10, width=70, state=DISABLED)
        self.log.pack(pady=10)
    
    def select_folder(self):
        folder = filedialog.askdirectory(title="Выберите папку с Excel-файлами")
        if folder:
            self.folder_path = folder
            self.lbl_folder.config(text=f"Папка: {folder}")
            self.btn_run.config(state=NORMAL)
    
    def log_message(self, message):
        """Добавляет сообщение в лог"""
        self.log.config(state=NORMAL)
        self.log.insert(END, message + "\n")
        self.log.see(END)
        self.log.config(state=DISABLED)
    
    def update_status(self, message):
        """Обновляет статус"""
        self.lbl_status.config(text=message)
        self.root.update()
    
    def run_processing(self):
        """Запускает обработку в отдельном потоке"""
        self.btn_run.config(state=DISABLED)
        self.btn_select.config(state=DISABLED)
        self.progress['value'] = 0
        
        # Очищаем лог
        self.log.config(state=NORMAL)
        self.log.delete(1.0, END)
        self.log.config(state=DISABLED)
        
        # Запускаем в отдельном потоке, чтобы GUI не зависал
        thread = threading.Thread(target=self.process_files)
        thread.start()
    
    def process_files(self):
        """Основная логика обработки файлов"""
        results = []
        result_json = {}
        counter = 0
        
        # Ищем все Excel-файлы
        self.update_status("Поиск Excel-файлов...")
        all_files = []
        for ext in ("*.xlsx", "*.xlsm", "*.xls"):
            all_files.extend(Path(self.folder_path).rglob(ext))
        
        if not all_files:
            self.update_status("Файлы не найдены")
            messagebox.showerror("Ошибка", "Excel-файлы не найдены в указанной папке")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        total_files = len(all_files)
        self.log_message(f"Найдено файлов: {total_files}")
        
        for i, filepath in enumerate(all_files):
            # Обновляем прогресс
            progress_value = (i + 1) / total_files * 100
            self.progress['value'] = progress_value
            
            # Показываем текущий файл
            self.update_status(f"Обработка: {filepath.name}")
            self.log_message(f"[{i+1}/{total_files}] {filepath.name}")
            
            try:
                wb = load_workbook(filepath, data_only=True, read_only=True)
            except Exception as e:
                self.log_message(f"  ✗ Ошибка открытия: {e}")
                continue
            
            # Проверяем наличие листа PLM_TEM_Report
            if "PLM_TEM_Report" not in wb.sheetnames:
                self.log_message(f"  ⊘ Лист PLM_TEM_Report не найден - пропуск")
                wb.close()
                continue
            
            # есть ли в репорте данные
            plm_result_sheet = wb["PLM_TEM_Report"]
            if plm_result_sheet.max_row < 6:
                self.log_message(f"  ⊘ Лист PLM_TEM_Report пустой")
                wb.close()
                continue

            counter += 1
            # 1. Берем информацию из M1
            project_info = plm_result_sheet["M1"].value
            date_analyzed = plm_result_sheet["M2"].value

            if project_info is not None:
                if str(project_info).strip() != '':
                    if str(project_info) not in str(filepath):
                        project_info = find_project_in_sample(str(plm_result_sheet["B6"].value))

            if project_info is None or str(project_info).strip() == '':
                project_info = str(plm_result_sheet["B6"].value)[:-2]


            print(project_info)
            if project_info in result_json:
                self.log_message(f" Already Has")
                wb.close()
                continue
            
            if project_info is None:
                continue
            
            if str(project_info) == 'No':
                continue
                
            if 'Conflict' in str(project_info):
                continue

            if str(project_info) in ['260407-106', '260417-4', '260508-22', '260514-17', '260525-18']:
                continue

            # 2. Берем информацию из B3 листа SampleAnalyses
            analyst_info = None
            if "SampleAnalyses" in wb.sheetnames:
                sample_sheet = wb["SampleAnalyses"]
                if sample_sheet["B3"].value:
                    if len(str(sample_sheet["B3"].value).strip()) > 1:
                        analyst_info = str(sample_sheet["B3"].value).strip().upper()
                else:
                    analyst_info = "No Analyst"
            else: 
                analyst_info = "No Analyst"
            
            # 3. Считаем строки
            plm_count = 0
            nob_count = 0
            tem_count = 0
            total_count = 0

            for row in plm_result_sheet.iter_rows(min_row=6, min_col=2, max_col=2):
                cell_i = row[0].value
                if cell_i is not None:
                    cell_str = str(cell_i).strip()
                    if (cell_str):
                        if (cell_str) and (project_info in cell_str):
                            total_count += 1

            for row in plm_result_sheet.iter_rows(min_row=6, min_col=9, max_col=17):
                # Колонка I (индекс 0, т.к. min_col=9)
                cell_i = row[0].value
                if cell_i is not None:
                    cell_str = str(cell_i).strip()
                    if cell_str and cell_str != "Not Applicable":
                        plm_count += 1
                
                # Колонка L (индекс 3, т.к. L=12, 12-9=3)
                cell_l = row[3].value
                if cell_l is not None:
                    cell_str = str(cell_l).strip()
                    if cell_str and cell_str != "Not Applicable":
                        nob_count += 1
                
                # Колонка Q (индекс 8, т.к. Q=17, 17-9=8)
                cell_q = row[8].value
                if cell_q is not None:
                    cell_str = str(cell_q).strip()
                    if cell_str and cell_str != "Not Analyzed":
                        tem_count += 1
            
            all_plm_data = []
            all_nob_data = []

            if "SampleAnalyses" in wb.sheetnames:
                plm_sample_sheet = wb["SampleAnalyses"]
                
                if plm_sample_sheet['A8'].value is not None:
                    for row in plm_sample_sheet.iter_rows(min_row=8, max_row=(8 + total_count - 1), min_col=1, max_col=46, values_only=True):
                        plm_analysis_json = {}
                        nob_analysis_json = {}
                        if row[0] is None:
                            continue
                        if str(row[43]).strip() == '198.1' or str(row[43]).strip() == '198,1':
                            for col in range(46):
                                # if row[col] is not None:
                                plm_analysis_json[plm_analysis_columns[col]] = str(row[col])
                        else:
                            for col in range(46):
                                # if row[col] is not None:
                                nob_analysis_json[plm_analysis_columns[col]] = str(row[col])

                        if  len(plm_analysis_json) > 0:
                            all_plm_data.append(plm_analysis_json)

                        if len(nob_analysis_json) > 0:
                            all_nob_data.append(nob_analysis_json)
                        

            all_tem_data = []
            if "TEM_Calculation" in wb.sheetnames:
                tem_sample_sheet = wb["TEM_Calculation"]

                if tem_sample_sheet['A6'].value is not None:
                    for row in tem_sample_sheet.iter_rows(min_row=6, max_row=(6 + total_count - 1), min_col=1, max_col=19, values_only=True):
                        tem_analysis_json = {}
                        if row[0] is None:
                            continue
                        elif row[18] == 'PS':
                            continue
                        else:
                            for col in range(19):
                                tem_analysis_json[tem_analysis_columns[col]] = str(row[col])
                        all_tem_data.append(tem_analysis_json)

            wb.close()
            
            if date_analyzed:
                pass
            else:
                date_analyzed = None

            result_json[project_info] = {"date": str(date_analyzed), 
                                         "analyst": analyst_info, 
                                         "plm_count": plm_count, 
                                         "nob_count": nob_count, 
                                         "tem_count": tem_count,
                                         "total_count": total_count,
                                          "file_name" : str(filepath), 
                                          "plm_analysis": all_plm_data,
                                          'nob_analysis': all_nob_data, 
                                          "tem_analysis": all_tem_data }
            
            output_file = Path(self.folder_path) / "qc_raw_data.json"

        with open(output_file, "a", encoding="utf-8") as f:
            f.write(json.dumps(result_json, indent=4, ensure_ascii=False))

        self.log_message(f"  ✓ PLM: {plm_count}, NOB: {nob_count}, TEM: {tem_count}")
        
        if not results:
            self.update_status("Готово (нет данных)")
            messagebox.showwarning("Внимание", "Не найдено файлов с листом PLM_TEM_Report")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        # Сохраняем результат
        self.update_status("Сохраняю файл...")
        self.log_message("\nСохраняю результат...")
        
        output_file = Path(self.folder_path) / "qc_data.xlsx"
    
        
        # headers = ["number", "project", "date", "analyst", "plm_count", "nob_count", "tem_count", "total_count", "file_name"]
        # sheet_raw.append(headers)
        # for r in results:
        #     sheet_raw.append([
        #         r["number"],
        #         r["project"],
        #         r["date"],
        #         r["analyst"],
        #         r["plm_count"],
        #         r["nob_count"],
        #         r["tem_count"],
        #         r["total_count"],
        #         str(r['file_name'])
        #     ])
        

        
        # result_wb.save(output_file)
      
        self.log_message("  ✓ Лист 'По аналитикам' создан")
        self.log_message("  ✓ Лист 'По месяцам' создан")
        self.log_message("  ✓ Лист 'Аналитик-Месяц' создан")
        
        self.progress['value'] = 100
        self.update_status("Готово!")
        self.log_message(f"\n✓ Сохранено: {output_file}")
        self.log_message(f"Обработано файлов с данными: {len(results)}")
        
        messagebox.showinfo(
            "Готово!", 
            f"Обработано файлов: {len(results)}\n\n")
        
        self.btn_run.config(state=NORMAL)
        self.btn_select.config(state=NORMAL)


if __name__ == "__main__":
    root = Tk()
    app = QCProcessorApp(root)
    root.mainloop()

    # pyinstaller --onefile --windowed --name="QC_Collector" qc_info0.07.py