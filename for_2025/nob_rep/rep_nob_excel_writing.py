import json
from pathlib import Path
import math
import random
import copy
import os
import logging
import sys
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from utils.utilities import calc_std_dev, return_excel_filename, replicate_analyst

logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s'
)

def main_start():
    script_dir = Path(__file__).resolve().parent
    excel_result_dir = script_dir.parent / "excel_result"

    for file in script_dir.glob("*.json"):
        file_base_name = os.path.basename(file)

        excel_result_name = return_excel_filename(type='NOB', input_filename=file_base_name)
        template_excel_path = Path(excel_result_dir / excel_result_name)
        print(template_excel_path)
        wb = load_workbook(template_excel_path)

        with file.open("r", encoding="utf-8") as f:
            duplicates_data = json.load(f)
        
        for analyst, info in duplicates_data.items():
            if analyst == 'total':
                ws = wb["Reps"]
                count = 4
                row_number = 1 
                for item_info in info:
                    text_date_for_excel = item_info['date'][5:7] + '-' + item_info['date'][8:10] + '-' + item_info['date'][0:4]
                    ws[f'A{count}'] = row_number
                    ws[f'B{count}'] = text_date_for_excel
                    ws[f'C{count}'] = item_info['project']
                    if '-1000' not in item_info['lab id']:
                        ws[f'D{count}'] = item_info['sample']
                        ws[f'E{count}'] = item_info['1st_analyst_name']
                        ws[f'F{count}'] = item_info['1st_analyst_asb_type'][:4] 
                        ws[f'G{count}'] = item_info['1st_analyst_asb_percent']
                        ws[f'H{count}'] = text_date_for_excel #
                        ws[f'I{count}'] = item_info['dup_name'] #
                        ws[f'J{count}'] = item_info['dup_asb_type'][:4] #
                        ws[f'K{count}'] = item_info['dup_asb_percent'] #
                        ws[f'L{count}'] = calc_std_dev(item_info['1st_analyst_asb_percent'], item_info['dup_asb_percent'])
                    else:
                        ws[f'D{count}'] = 'BL'
                        ws[f'E{count}'] = item_info['1st_analyst_name']
                        ws[f'F{count}'] = 'NAD'
                        ws[f'G{count}'] = "NAD"
                        ws[f'H{count}'] = text_date_for_excel #
                        ws[f'I{count}'] = item_info['dup_name'] #
                        ws[f'J{count}'] = 'NAD'
                        ws[f'K{count}'] = 'NAD'
                        ws[f'L{count}'] = 0
                    count += 1
                    row_number += 1
            else:
                ws = wb[analyst]
                count = 17
                for item_info in info:
                    logging.info(item_info['project'])
                    if '-1000' not in item_info['lab id']:
                        ws[f'J{count}'] = item_info['1st_analyst_name']
                        ws[f'K{count}'] = item_info['1st_analyst_asb_type'][:4] 
                        ws[f'L{count}'] = item_info['1st_analyst_asb_percent'] #
                        ws[f'M{count}'] = item_info['dup_name']
                        ws[f'N{count}'] = item_info['dup_asb_type'][:4]
                        ws[f'O{count}'] = item_info['dup_asb_percent']
                        ws[f'P{count}'] = calc_std_dev(item_info['1st_analyst_asb_percent'], item_info['dup_asb_percent'])
                    else:
                        ws[f'J{count}'] = analyst
                        ws[f'K{count}'] = 'NAD'
                        ws[f'L{count}'] = 'NAD'
                        ws[f'M{count}'] = replicate_analyst(analyst)
                        ws[f'N{count}'] = 'NAD'
                        ws[f'O{count}'] = 'NAD'
                        ws[f'P{count}'] = 0
                    count += 1

        wb.save(template_excel_path)

if __name__ == "__main__":
    main_start()