import threading
from pathlib import Path
from tkinter import Tk, Button, Label, Text, filedialog, messagebox, END, DISABLED, NORMAL
from tkinter.ttk import Progressbar
from openpyxl import Workbook, load_workbook
import pandas as pd
from datetime import date
from collections import Counter

def sort_key(item):
    project = item['project']
    parts = project.split('-')
    return (parts[0], int(parts[1]))  # дата как строка, номер как число


class QCProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("QC Processor")
        self.root.geometry("600x400")
        self.root.resizable(False, False)
        
        self.folder_path = None
        
        # Кнопка выбора папки
        self.btn_select = Button(
            root, 
            text="Choose Folder", 
            command=self.select_folder,
            width=20,
            height=2
        )
        self.btn_select.pack(pady=10)
        
        # Метка с выбранной папкой
        self.lbl_folder = Label(root, text="Папка не выбрана", wraplength=550)
        self.lbl_folder.pack(pady=5)
        
        # Кнопка запуска
        self.btn_run = Button(
            root, 
            text="Запустить", 
            command=self.run_processing,
            width=20,
            height=2,
            state=DISABLED
        )
        self.btn_run.pack(pady=10)
        
        # Прогресс-бар
        self.progress = Progressbar(root, length=550, mode='determinate')
        self.progress.pack(pady=10)
        
        # Текущий статус
        self.lbl_status = Label(root, text="", font=("Arial", 10))
        self.lbl_status.pack(pady=5)
        
        # Лог
        self.log = Text(root, height=10, width=70, state=DISABLED)
        self.log.pack(pady=10)
    
    def select_folder(self):
        folder = filedialog.askdirectory(title="Выберите папку с Excel-файлами")
        if folder:
            self.folder_path = folder
            self.lbl_folder.config(text=f"Папка: {folder}")
            self.btn_run.config(state=NORMAL)
    
    def log_message(self, message):
        """Добавляет сообщение в лог"""
        self.log.config(state=NORMAL)
        self.log.insert(END, message + "\n")
        self.log.see(END)
        self.log.config(state=DISABLED)
    
    def update_status(self, message):
        """Обновляет статус"""
        self.lbl_status.config(text=message)
        self.root.update()
    
    def run_processing(self):
        """Запускает обработку в отдельном потоке"""
        self.btn_run.config(state=DISABLED)
        self.btn_select.config(state=DISABLED)
        self.progress['value'] = 0
        
        # Очищаем лог
        self.log.config(state=NORMAL)
        self.log.delete(1.0, END)
        self.log.config(state=DISABLED)
        
        # Запускаем в отдельном потоке, чтобы GUI не зависал
        thread = threading.Thread(target=self.process_files)
        thread.start()
    
    def process_files(self):
        """Основная логика обработки файлов"""
        results = []
        counter = 0
        
        # Ищем все Excel-файлы
        self.update_status("Поиск Excel-файлов...")
        all_files = []
        for ext in ("*.xlsx", "*.xlsm", "*.xls"):
            all_files.extend(Path(self.folder_path).rglob(ext))
        
        if not all_files:
            self.update_status("Файлы не найдены")
            messagebox.showerror("Ошибка", "Excel-файлы не найдены в указанной папке")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        total_files = len(all_files)
        self.log_message(f"Найдено файлов: {total_files}")
        
        for i, filepath in enumerate(all_files):
            # Обновляем прогресс
            progress_value = (i + 1) / total_files * 100
            self.progress['value'] = progress_value
            
            # Показываем текущий файл
            self.update_status(f"Обработка: {filepath.name}")
            self.log_message(f"[{i+1}/{total_files}] {filepath.name}")
            
            try:
                wb = load_workbook(filepath, data_only=True, read_only=True)
            except Exception as e:
                self.log_message(f"  ✗ Ошибка открытия: {e}")
                continue
            
            # Проверяем наличие листа PLM_TEM_Report
            if "PLM_TEM_Report" not in wb.sheetnames:
                self.log_message(f"  ⊘ Лист PLM_TEM_Report не найден - пропуск")
                wb.close()
                continue
            
            # есть ли в репорте данные
            plm_sheet = wb["PLM_TEM_Report"]
            if plm_sheet.max_row < 6:
                self.log_message(f"  ⊘ Лист PLM_TEM_Report пустой")
                wb.close()
                continue
                
            counter += 1
            # 1. Берем информацию из M1
            project_info = plm_sheet["M1"].value
            date_analyzed = plm_sheet["M2"].value

            
            # 2. Берем информацию из B3 листа SampleAnalyses
            analyst = None
            if "SampleAnalyses" in wb.sheetnames:
                sample_sheet = wb["SampleAnalyses"]
                if sample_sheet["B3"].value:
                    if len(str(sample_sheet["B3"].value).strip()) > 1:
                        analyst = str(sample_sheet["B3"].value).strip().upper()
                else:
                    analyst = "No Analyst"
            else: 
                analyst = "No Analyst"
            
            nob_calc_sheet = wb["NOB_Calculation"]

            # 3. Считаем строки
            bl_count = 0
            qc_count = 0

            for row in nob_calc_sheet.iter_rows(min_row=6, min_col=1, max_col=2):
                # Колонка I (индекс 0, т.к. min_col=9)
                cell_i = row[0].value
                if cell_i is not None:
                    cell_str = str(cell_i).strip().lower()
                    if cell_str and cell_str == "bl":
                        bl_count += 1

            sample_max_col = 0
            iter = 7
            while True: 
                # print(sample_sheet[f'B{iter}'].value)
                if sample_sheet[f'B{iter}'].value == "" or sample_sheet[f'B{iter}'].value == None: 
                    sample_max_col = iter
                    break
                iter+=1


            values = []
            for row in range(8, sample_max_col + 1):
                cell_value = sample_sheet[f'B{row}'].value
                if cell_value is not None:
                    values.append(cell_value)

            # 2. Считаем сколько раз встречается каждое значение
            counter_dup = Counter(values)
            
            # 3. Считаем дубликаты
            total_duplicates = 0
            for value, count in counter_dup.items():
                if count > 1:
                    total_duplicates += count - 1 

            qc_count = total_duplicates

            wb.close()

            results.append({
                "number": counter,
                "project": project_info,
                "analyst": analyst,
                "bl_count": bl_count,
                "qc_count": qc_count
            })
            
            self.log_message(f"  ✓ BL: {bl_count}, QC: {qc_count}")
        
        if not results:
            self.update_status("Готово (нет данных)")
            messagebox.showwarning("Внимание", "Не найдено файлов с листом PLM_TEM_Report")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        # Сохраняем результат
        self.update_status("Сохраняю файл...")
        self.log_message("\nСохраняю результат...")
        
        output_file = Path(self.folder_path) / "qc_duplicates.xlsx"
        
        result_wb = Workbook()
        sheet_raw = result_wb.active
        sheet_raw.title = "QC Duplicates"
        
        headers = ["number", "project", "analyst", "bl_count", "qc_count"]
        sheet_raw.append(headers)
        
        results = sorted(results, key=sort_key)

        final_raw_num = 1
        for r in results:
            sheet_raw.append([
                final_raw_num,
                r["project"],
                r["analyst"],
                r["bl_count"],
                r["qc_count"],
            ])
            final_raw_num += 1
        
        final_raw_num = 1

        result_wb.save(output_file)

        raw_df = pd.DataFrame(results, columns=["number", "project", "analyst", "bl_count", "qc_count" ])
                                           
        #  ws_raw.append([number, row["project"], row["date"],row["analyst"],row["plm_count"],row["nob_count"],row["tem_count"],row["month"],]),

        for analyst in raw_df['analyst'].unique():
            analyst_df = raw_df[raw_df['analyst'] == analyst]
            ws_analyst = result_wb.create_sheet(title=f'{analyst}_dupl')
            ws_analyst.append(["number", "project", "analyst", "bl_count", "qc_count"])

            for item, row in analyst_df.iterrows():
                ws_analyst.append([final_raw_num, row["project"], row["analyst"],row["bl_count"],row["qc_count"]])
                final_raw_num += 1


        # result_wb.save(output_file)


        # result_wb = load_workbook(output_file)
        # if 'QC Raw Data' in result_wb.sheetnames:
        #     print('test')
        #     sheet_to_delete = result_wb['QC Raw Data']
        #     result_wb.remove(sheet_to_delete)

        
        result_wb.save(output_file)
        
        self.log_message("  ✓ Лист 'По аналитикам' создан")
        self.log_message("  ✓ Лист 'По месяцам' создан")
        self.log_message("  ✓ Лист 'Аналитик-Месяц' создан")
        
        self.progress['value'] = 100
        self.update_status("Готово!")
        self.log_message(f"\n✓ Сохранено: {output_file}")
        self.log_message(f"Обработано файлов с данными: {len(results)}")
        
        messagebox.showinfo(
            "Готово!", 
            f"Обработано файлов: {len(results)}\n\n")
        
        self.btn_run.config(state=NORMAL)
        self.btn_select.config(state=NORMAL)


if __name__ == "__main__":
    root = Tk()
    app = QCProcessorApp(root)
    root.mainloop()

    # pyinstaller --onefile --windowed --name="QC_Collector" qc_info0.07.py