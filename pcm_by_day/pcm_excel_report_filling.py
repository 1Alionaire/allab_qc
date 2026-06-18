import json
from pathlib import Path
import math
import random
import copy
import os
import logging
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime, timedelta
import logging, traceback, os, sys

def get_writable_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)   # папка рядом с .exe
    return os.path.dirname(os.path.abspath(__file__))

script_dir = Path(__file__).resolve().parent
date = None

types_analysis = ['plm', 'nob', 'tem']

def generate_report_excels(input_data):
    wb = Workbook()
    unique_analysts = {value["analyst"] for value in input_data.values()}

    date_for_excel_file_raw = list(input_data.items())[0][1]['date_analyzed']

    date_for_excel_file = date_for_excel_file_raw[5:7] + '-' + date_for_excel_file_raw[8:10] + '-' + date_for_excel_file_raw[0:4]
    for analyst in unique_analysts:
        selected_analyst_ws = wb.create_sheet(f"{analyst}")
        selected_analyst_data = {key: value for key, value in input_data.items() if value['analyst'] == analyst }

        selected_analyst_ws['A1'] = 'Lab ID'
        selected_analyst_ws['B1'] = 'Date Analyzed'
        selected_analyst_ws['C1'] = 'Analyst'
        selected_analyst_ws['D1'] = 'Original Value'
        selected_analyst_ws['E1'] = 'QC Value'

        row = 2
        for key, value in selected_analyst_data.items():
            selected_analyst_ws[f'A{row}'] = key
            selected_analyst_ws[f'B{row}'] = value['date_analyzed'][5:7] + '-' + value['date_analyzed'][8:10] + '-' + value['date_analyzed'][0:4]
            selected_analyst_ws[f'c{row}'] = value['analyst']
            selected_analyst_ws[f'D{row}'] = value['original_value']
            selected_analyst_ws[f'E{row}'] = value['qc_value']
            row += 1

    out_path = os.path.join(get_writable_dir(),  f"PCM_{date_for_excel_file}.xlsx")
    wb.save(out_path)


# if __name__ == '__main__':
#     file = script_dir / 'qc_result.json'

#     with file.open("r", encoding="utf-8") as f:
#         duplicates_data = json.load(f)
    
#     generate_report_excels(duplicates_data)

    
