import threading
from pathlib import Path
from tkinter import Tk, Button, Label, Text, filedialog, messagebox, END, DISABLED, NORMAL
from tkinter.ttk import Progressbar
from openpyxl import Workbook, load_workbook
import pandas as pd
from datetime import date
import json
import time

def find_project_in_sample(inp_sample_id):
    final_result = ''
    count = 2
    for i in inp_sample_id:
        if count == 0:
            return final_result[:-1]
        final_result += i
        if i == '-':
            count -= 1

def normalize_value(input_value):
    inp_str = str(input_value).strip()
    if '..' in inp_str:
        inp_str = inp_str.replace('..', '.')
    if ',' in inp_str:
        inp_str = inp_str.replace(',', '.')
    return float(inp_str)

def random_calc_point(original_value):
    operation = random.choice(['-', '+'])
    if operation == '-':
        value = random.choice([1, 2, 3])
        if (float(original_value) - value) > 0:
            return float(original_value) - value
        else:
            return 2.0
    else:
        value = random.choice([1, 2, 3])
        return float(inp_str) + value

class QCProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("QC Processor")
        self.root.geometry("600x400")
        self.root.resizable(False, False)
        
        self.folder_path = None
        
        # Кнопка выбора папки
        self.btn_select = Button(
            root, 
            text="Choose Folder", 
            command=self.select_folder,
            width=20,
            height=2
        )
        self.btn_select.pack(pady=10)
        
        # Метка с выбранной папкой
        self.lbl_folder = Label(root, text="Папка не выбрана", wraplength=550)
        self.lbl_folder.pack(pady=5)
        
        # Кнопка запуска
        self.btn_run = Button(
            root, 
            text="Запустить", 
            command=self.run_processing,
            width=20,
            height=2,
            state=DISABLED
        )
        self.btn_run.pack(pady=10)
        
        # Прогресс-бар
        self.progress = Progressbar(root, length=550, mode='determinate')
        self.progress.pack(pady=10)
        
        # Текущий статус
        self.lbl_status = Label(root, text="", font=("Arial", 10))
        self.lbl_status.pack(pady=5)
        
        # Лог
        self.log = Text(root, height=10, width=70, state=DISABLED)
        self.log.pack(pady=10)
    
    def select_folder(self):
        folder = filedialog.askdirectory(title="Выберите папку с Excel-файлами")
        if folder:
            self.folder_path = folder
            self.lbl_folder.config(text=f"Папка: {folder}")
            self.btn_run.config(state=NORMAL)
    
    def log_message(self, message):
        """Добавляет сообщение в лог"""
        self.log.config(state=NORMAL)
        self.log.insert(END, message + "\n")
        self.log.see(END)
        self.log.config(state=DISABLED)
    
    def update_status(self, message):
        """Обновляет статус"""
        self.lbl_status.config(text=message)
        self.root.update()
    
    def run_processing(self):
        """Запускает обработку в отдельном потоке"""
        self.btn_run.config(state=DISABLED)
        self.btn_select.config(state=DISABLED)
        self.progress['value'] = 0
        
        # Очищаем лог
        self.log.config(state=NORMAL)
        self.log.delete(1.0, END)
        self.log.config(state=DISABLED)
        
        # Запускаем в отдельном потоке, чтобы GUI не зависал
        thread = threading.Thread(target=self.process_files)
        thread.start()
    
    def process_files(self):
        """Основная логика обработки файлов"""
        results = []
        pcm_sample_dict = {}
        counter = 0
        pcm_sample_array = []
        
        # Ищем все Excel-файлы
        self.update_status("Поиск Excel-файлов...")
        all_files = []
        for ext in ("*.xlsx", "*.xlsm", "*.xls"):
            all_files.extend(Path(self.folder_path).rglob(ext))
        
        if not all_files:
            self.update_status("Файлы не найдены")
            messagebox.showerror("Ошибка", "Excel-файлы не найдены в указанной папке")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        total_files = len(all_files)
        self.log_message(f"Найдено файлов: {total_files}")
        
        for i, filepath in enumerate(all_files):
            # Обновляем прогресс
            progress_value = (i + 1) / total_files * 100
            self.progress['value'] = progress_value
            
            # Показываем текущий файл
            self.update_status(f"Обработка: {filepath.name}")
            self.log_message(f"[{i+1}/{total_files}] {filepath.name}")
            
            try:
                wb = load_workbook(filepath, data_only=True, read_only=True)
            except Exception as e:
                self.log_message(f"  ✗ Ошибка открытия: {e}")
                continue
            
            # Проверяем наличие листа PLM_TEM_Report
            if "Count-Recount" not in wb.sheetnames:
                self.log_message(f"  ⊘ Лист PCM не найден - пропуск")
                wb.close()
                continue
            
            if 'Conflict' in str(filepath):
                self.log_message(f" Duplicate Conflict")
                wb.close()
                continue

            self.log_message("*" * 50)
            
            pcm_qc_sheet = wb["Count-Recount"]
            counter += 1
            # 1. Берем информацию из M1
            project_info = pcm_qc_sheet["B2"].value
            date_analyzed = pcm_qc_sheet["L3"].value

            if project_info is not None:
                if str(project_info).strip() != '':
                    if str(project_info) not in str(filepath):
                        project_info = find_project_in_sample(wb["Sample"]["Q5"].value)

            self.log_message(f'project_info: {project_info}')
            
            row_count = 0
            client_sample_id = None
            original_sample_value = 0
            qc_sample_value = 0
            self.log_message('*' * 50)
            self.log_message(f" project_info: {project_info} ")

            last_sample_row = 0
            for i in range(8, 46):
                if str(pcm_qc_sheet[f"B{i}"].value).strip().lower() == 'overload':
                    continue
                if str(pcm_qc_sheet[f"B{i}"].value) == 'None':
                    last_sample_row = i - 1
                    break
                if float(pcm_qc_sheet[f"B{i}"].value) <= 0.5:
                    last_sample_row = i - 1
                    break

            self.log_message(f'last_sample_row: {last_sample_row}')

            have_sample = False
            for i in range(8, last_sample_row):
                row_count += 1
                if (str(pcm_qc_sheet[f"F{i}"].value) != 'None' 
                    and str(pcm_qc_sheet[f"G{i}"].value) != 'None'
                    and str(pcm_qc_sheet[f"H{i}"].value) != 'None'):
                        
                        have_sample = True
                        if str(pcm_qc_sheet[f"A{i}"].value) != '0':
                            client_sample_id = str(project_info) + '-' + str(row_count)

                            if client_sample_id in pcm_sample_dict:
                                continue

                            qc_sample_value = normalize_value(pcm_qc_sheet[f"F{i}"].value)
                            original_sample_value = normalize_value(pcm_qc_sheet[f"B{i}"].value)

                            pcm_sample_dict[client_sample_id] = {'original_value' : original_sample_value,
                                                                        'qc_value' : qc_sample_value,
                                                                        'analyst' : str(pcm_qc_sheet["S3"].value), 
                                                                        'date_analyzed' : str(date_analyzed) }
            
            if have_sample == False:
                self.log_message(f'have_sample: {have_sample}')
                if (last_sample_row - 7) < 4:
                    self.log_message(f'amount_samples: {(last_sample_row - 7)}')
                    pass
                else:
                    pcm_sample_dict[str(project_info) + '-1'] = {'error' : 'No sample'}
                    
            wb.close()
            time.sleep(0.5) 
        output_file = Path(self.folder_path) / "qc_pcm_raw_data.json"

        with open(output_file, "a", encoding="utf-8") as f:
            f.write(json.dumps(pcm_sample_dict, indent=4, ensure_ascii=False))

        self.log_message(f" PCM Complete ✓ ")
        
        if not results:
            self.update_status("Готово (нет данных)")
            messagebox.showwarning("Внимание", "Не найдено файлов с листом PLM_TEM_Report")
            self.btn_run.config(state=NORMAL)
            self.btn_select.config(state=NORMAL)
            return
        
        # Сохраняем результат
        self.update_status("Сохраняю файл...")
        self.log_message("\nСохраняю результат...")
        
        output_file = Path(self.folder_path) / "qc_data.xlsx"
        
        self.progress['value'] = 100
        self.update_status("Готово!")
        self.log_message(f"\n✓ Сохранено: {output_file}")
        self.log_message(f"Обработано файлов с данными: {len(results)}")
        
        messagebox.showinfo(
            "Готово!", 
            f"Обработано файлов: {len(results)}\n\n")
        
        self.btn_run.config(state=NORMAL)
        self.btn_select.config(state=NORMAL)


if __name__ == "__main__":
    root = Tk()
    app = QCProcessorApp(root)
    root.mainloop()

    # pyinstaller --onefile --windowed --name="QC_PCM_Collector" parsing_all_pcm_files.py