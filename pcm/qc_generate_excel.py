from pathlib import Path
import json
import re
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time

INPUT_FILE = Path("qc_pcm_raw_data.json")
OUTPUT_FOLDER = Path("monthly_reports")

OUTPUT_FOLDER.mkdir(exist_ok=True)


def parse_date(value):
    """
    Превращает строку типа '2025-08-01 00:00:00' в datetime.
    """
    return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
    # return datetime.fromisoformat(value)


def clean_sheet_name(name):
    """
    Excel не разрешает некоторые символы в названии листа.
    Также название листа максимум 31 символ.
    """
    name = str(name).strip()
    name = re.sub(r'[\[\]\:\*\?\/\\]', "_", name)

    if not name:
        name = "Unknown"

    return name[:31]


def make_unique_sheet_name(name, used_names):
    """
    Если после очистки названия совпали, делаем уникальное имя:
    Analyst
    Analyst_2
    Analyst_3
    """
    base_name = clean_sheet_name(name)
    sheet_name = base_name
    counter = 2

    while sheet_name in used_names:
        suffix = f"_{counter}"
        sheet_name = base_name[:31 - len(suffix)] + suffix
        counter += 1

    used_names.add(sheet_name)
    return sheet_name


def lab_id_sort_key(lab_id):
    """
    Чтобы сортировка lab id была нормальная:

    250801-2
    250801-10

    а не:

    250801-10
    250801-2
    """
    result = []

    for part in str(lab_id).split("-"):
        if part.isdigit():
            result.append(int(part))
        else:
            result.append(part)

    return result


def style_sheet(ws):
    """
    Немного форматирования для каждого листа.
    """
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)

        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)


with INPUT_FILE.open("r", encoding="utf-8") as file:
    raw_data = json.load(file)


# month -> analyst -> rows
grouped_data = defaultdict(lambda: defaultdict(list))

for lab_id, info in raw_data.items():
    print(lab_id)
    print(info["date_analyzed"])
    date_analyzed = parse_date(info["date_analyzed"])

    # Важно: месяц берём именно из date_analyzed, а не из lab_id
    month_key = date_analyzed.strftime("%Y-%m")

    analyst = info.get("analyst") or "Unknown"

    row = {
        "lab_id": lab_id,
        "date_analyzed": date_analyzed,
        "analyst": analyst,
        "original_value": info.get("original_value"),
        "qc_value": info.get("qc_value"),
        # "low_range_sr": info.get("low_range_sr"),
    }

    grouped_data[month_key][analyst].append(row)


headers = [
    "lab_id",
    "date_analyzed",
    "analyst",
    "original_value",
    "qc_value",
    # "low_range_sr",
]


for month_key, analysts_dict in sorted(grouped_data.items()):
    wb = Workbook()

    # Удаляем дефолтный лист Sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    used_sheet_names = set()

    for analyst, rows in sorted(analysts_dict.items()):
        sheet_name = make_unique_sheet_name(analyst, used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)

        ws.append(headers)

        # Сортировка внутри листа:
        # сначала по дате анализа, потом по lab_id
        rows.sort(
            key=lambda row: (
                row["date_analyzed"],
                lab_id_sort_key(row["lab_id"]),
            )
        )

        for row in rows:
            ws.append([
                row["lab_id"],
                row["date_analyzed"],
                row["analyst"],
                row["original_value"],
                row["qc_value"],
                # row["low_range_sr"],
            ])

        # Формат даты в колонке B
        for cell in ws["B"][1:]:
            cell.number_format = "yyyy-mm-dd"

        style_sheet(ws)

    output_file = OUTPUT_FOLDER / f"qc_pcm_{month_key}.xlsx"
    wb.save(output_file)
    wb.close()
    time.sleep(0.5)

    print(f"Saved: {output_file}")

    # pyinstaller --onefile --windowed --name="QC_Generate_PCM_Count_Recount" qc_generate_excel.py