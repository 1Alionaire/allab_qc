import json
from pathlib import Path
import math
import random
import copy
import os
import sys
from openpyxl import load_workbook

path = Path('/Users/alibekkubeev/Downloads/for alibek')

for file in path.glob("*.xlsm"):
    wb = load_workbook(file)
    result_ws = wb['PLM_TEM_Report']
    for row in result_ws.iter_rows(min_row=6, min_col=9, max_col=17):
        cell_q = row[8].value
        if cell_q is not None:
            cell_str = str(cell_q).strip()
            if cell_str:
                if 'Chry' in cell_str:
                    print(file)
                    break

